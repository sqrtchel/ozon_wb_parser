# ============================================
# ПАРСЕР WILDBERRIES - ПОЛНАЯ ВЕРСИЯ (ДОЗАПИСЬ В БД С ДИАГНОСТИКОЙ)
# ============================================

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import json
import re
import time
import random
from datetime import datetime
import psycopg2
from psycopg2 import sql
import logging
from typing import Dict, Optional, Tuple
import os
import urllib.request
from openpyxl import Workbook, load_workbook

# ============================================
# НАСТРОЙКИ ЛОГГИРОВАНИЯ
# ============================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'wb_parser_{datetime.now().strftime("%Y%m%d")}.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================
# НАСТРОЙКИ ЗАДЕРЖЕК
# ============================================
MIN_DELAY = 1.5
MAX_DELAY = 3.0

# ============================================
# НАСТРОЙКИ БАЗЫ ДАННЫХ (ЗАМЕНИТЕ НА СВОИ)
# ============================================
DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'database': 'postgres',
    'user': 'postgres',
    'password': '1'
}

DB_SCHEMA = 'public'
TABLE_NAME = 'wildberries'
REFERENCE_TABLE = 'reference_wildberries'
EXCEL_FILE = "results.xlsx"
CHROME_DEBUG_ADDRESS = os.environ.get("CHROME_DEBUG_ADDRESS", "127.0.0.1:9222")


# ============================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================
def is_debug_chrome_ready(address: str = CHROME_DEBUG_ADDRESS) -> bool:
    """Проверяет доступность Chrome DevTools endpoint."""
    try:
        urllib.request.urlopen(f"http://{address}/json/version", timeout=2.0)
        return True
    except Exception:
        return False


def random_delay():
    """Случайная задержка для имитации человеческого поведения"""
    delay = random.uniform(MIN_DELAY, MAX_DELAY)
    logger.info(f"⏱️ Пауза {delay:.1f} сек...")
    time.sleep(delay)


def short_delay():
    """Короткая задержка"""
    delay = random.uniform(0.5, 1.0)
    time.sleep(delay)


def extract_number(text: str) -> Optional[float]:
    """Извлечение числа из текста"""
    numbers = re.findall(r'[\d\s]+', text)
    if numbers:
        try:
            return float(numbers[0].replace(' ', ''))
        except:
            return None
    return 0


# ============================================
# НАСТРОЙКА БРАУЗЕРА
# ============================================
def setup_driver(headless: bool = False):
    """
    Подключает Selenium к уже запущенному Chrome через remote debugging.
    Chrome должен быть открыт заранее пользователем.
    """
    logger.info("🚀 Подключение к открытому Chrome (debug mode)...")

    if not is_debug_chrome_ready(CHROME_DEBUG_ADDRESS):
        raise RuntimeError(
            f"Chrome debug endpoint недоступен: {CHROME_DEBUG_ADDRESS}.\n"
            "Перед запуском wb_pars.py выполните шаги:\n"
            "1) taskkill /F /IM chrome.exe /T\n"
            "2) Запустите Chrome с --remote-debugging-port=9222 и отдельным user-data-dir\n"
            "3) Проверьте endpoint командой Invoke-WebRequest http://127.0.0.1:9222/json/version\n"
            "4) Убедитесь, что mpstats установлен и авторизован в этом профиле."
        )

    chrome_options = Options()
    chrome_options.debugger_address = CHROME_DEBUG_ADDRESS

    if headless:
        logger.warning("⚠️ Headless режим игнорируется при подключении к debug Chrome.")

    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(40)
    logger.info(f"✅ Подключение к Chrome успешно: {CHROME_DEBUG_ADDRESS}")
    random_delay()
    return driver


# ============================================
# ФУНКЦИИ ПОИСКА ДАННЫХ
# ============================================
def find_old_price(driver) -> Optional[float]:
    """Поиск зачеркнутой цены"""
    print("\n💰 Шаг 1: Поиск зачеркнутой цены...")
    random_delay()

    old_price_selectors = [
        "span.productLinePriceOld--M0lnS",
        "span.price-block__old-price",
        "del.price-block__old-price",
        "span.mo-typography_color_secondary.price-block__old-price",
        "[class*='old-price']",
        "[class*='Old']",
        "ins.price-block__old-price",
        "span.price-old"
    ]

    for selector in old_price_selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            for element in elements:
                text = element.text.strip()
                if text and '₽' in text:
                    price = extract_number(text)
                    if price:
                        print(f"✅ Найдено через селектор: {selector}")
                        print(f"✅ Зачеркнутая цена: {price:.2f} ₽")
                        return price
            short_delay()
        except:
            continue

    # Пробуем через JavaScript
    print("🔄 Пробуем найти через JavaScript...")
    try:
        old_price_js = driver.execute_script("""
            const elements = document.querySelectorAll('span, del, ins');
            for (let el of elements) {
                const text = el.innerText;
                const isStrikethrough = window.getComputedStyle(el).textDecoration.includes('line-through');
                if (isStrikethrough && text.includes('₽')) {
                    const match = text.match(/(\\d[\\d\\s]*)/);
                    if (match) return match[1].replace(/\\s/g, '');
                }
            }
            return null;
        """)
        if old_price_js:
            price = float(old_price_js)
            print(f"✅ Зачеркнутая цена (JS): {price:.2f} ₽")
            return price
    except:
        pass

    print("⚠️ Зачеркнутая цена не найдена")
    return 0


def find_current_price(driver) -> Optional[float]:
    """Поиск обычной (текущей) цены."""
    print("\n💵 Шаг 2: Поиск обычной цены...")
    random_delay()

    current_price_selectors = [
        "ins.mo-typography.mo-typography_variant_body.mo-typography_variable-weight_body.mo-typography_variable.mo-typography_color_primary.priceBlockFinalPrice--iToZR",
        "ins.priceBlockFinalPrice--iToZR",
        "ins[class*='priceBlockFinalPrice']",
        "ins[class*='FinalPrice']",
    ]

    for selector in current_price_selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            for element in elements:
                text = element.text.strip()
                if text and "₽" in text:
                    price = extract_number(text)
                    if price:
                        print(f"✅ Обычная цена найдена через: {selector}")
                        print(f"💵 Обычная цена: {price:.2f} ₽")
                        return price
            short_delay()
        except:
            continue

    print("⚠️ Обычная цена не найдена")
    return 0


def find_product_name(driver) -> str:
    """Поиск наименования товара"""
    print("\n📝 Шаг 3: Поиск наименования...")
    random_delay()

    name_selectors = [
        "h2.productTitle--lfc4o",
        "h1.product-page__title",
        "h1.product-title",
        "h1[class*='title']",
        ".product-name"
    ]

    for selector in name_selectors:
        try:
            element = driver.find_element(By.CSS_SELECTOR, selector)
            name = element.text.strip()
            if name:
                print(f"✅ Наименование найдено через: {selector}")
                print(f"📝 Наименование: {name}")
                return name
        except:
            continue

    # Пробуем через title страницы
    try:
        title = driver.title
        if title and ' - купить' in title:
            name = title.split(' - купить')[0]
            if name:
                print(f"✅ Наименование найдено через title")
                print(f"📝 Наименование: {name}")
                return name
    except:
        pass

    print("⚠️ Наименование не найдено")
    return "Не найдено"


def find_article_number(driver) -> Tuple[int, str]:
    """Поиск артикула несколькими способами"""
    print("\n🔢 Шаг 4: Поиск артикула...")
    random_delay()

    # Способ 1: Через селектор из URL
    try:
        current_url = driver.current_url
        url_match = re.search(r'/catalog/(\d+)/', current_url)
        if url_match:
            article = int(url_match.group(1))
            print(f"✅ Артикул найден в URL: {article}")
            return article, "URL"
    except:
        pass

    # Способ 2: Различные CSS селекторы
    article_selectors = [
        "span.nominal-article__article",
        "span.article",
        "div.product-article span",
        "span[class*='article']",
        "span[data-link='article']",
        "td.cellValue--hHBJB button span",
        "span.product-article__value",
        "div.product-article__article",
        "span[data-tag='article']"
    ]

    for selector in article_selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            for element in elements:
                text = element.text.strip()
                text = re.sub(r'[^0-9]', '', text)
                if text and text.isdigit():
                    article = int(text)
                    print(f"✅ Артикул найден через селектор: {selector}")
                    print(f"   Значение: {article}")
                    return article, selector
            short_delay()
        except:
            continue

    # Способ 3: Поиск через JavaScript
    try:
        article_js = driver.execute_script("""
            const elements = document.querySelectorAll('span, div, p, td');
            for (let el of elements) {
                const text = el.innerText;
                if (text && text.includes('Артикул')) {
                    const nextEl = el.nextElementSibling;
                    if (nextEl) {
                        const numbers = nextEl.innerText.match(/\\d+/);
                        if (numbers) return numbers[0];
                    }
                    const numbers = text.match(/\\d+/);
                    if (numbers) return numbers[0];
                }
            }
            return null;
        """)

        if article_js:
            article = int(article_js)
            print(f"✅ Артикул найден через JavaScript: {article}")
            return article, "JavaScript"
    except:
        pass

    # Способ 4: Из JSON-LD данных
    try:
        script_elements = driver.find_elements(By.CSS_SELECTOR, 'script[type="application/ld+json"]')
        for script in script_elements:
            try:
                data = json.loads(script.get_attribute('innerHTML'))
                if 'sku' in data:
                    article = int(data['sku'])
                    print(f"✅ Артикул найден в JSON-LD: {article}")
                    return article, "JSON-LD"
                elif 'productID' in data:
                    article = int(data['productID'])
                    print(f"✅ Артикул найден в JSON-LD (productID): {article}")
                    return article, "JSON-LD"
            except:
                continue
    except:
        pass

    # Способ 5: Из meta-тегов
    try:
        meta_article = driver.find_element(By.CSS_SELECTOR, 'meta[name="article"]')
        if meta_article:
            article = int(meta_article.get_attribute('content'))
            print(f"✅ Артикул найден в meta-теге: {article}")
            return article, "meta"
    except:
        pass

    # Способ 6: Из data-атрибутов
    try:
        data_article = driver.find_element(By.CSS_SELECTOR, '[data-article]')
        if data_article:
            article = int(data_article.get_attribute('data-article'))
            print(f"✅ Артикул найден в data-атрибуте: {article}")
            return article, "data-attribute"
    except:
        pass

    print("⚠️ Артикул не найден ни одним способом")
    return 0, "not_found"


def find_sells(driver) -> Optional[float]:
    """Поиск продаж по блоку вида '27 шт.'."""
    print("\n📊 Шаг 5: Поиск продаж...")
    random_delay()

    primary_selector = "span._text_j8xb1_1._body-md-medium_j8xb1_75._default_j8xb1_109"
    sells_values = []

    try:
        elements = driver.find_elements(By.CSS_SELECTOR, primary_selector)
        for element in elements:
            text = element.text.strip().lower()
            if "шт" not in text:
                continue

            match = re.search(r'(\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*шт\.?', text)
            if not match:
                continue

            normalized = match.group(1).replace(' ', '').replace('\u00a0', '').replace(',', '.')
            try:
                value = float(normalized)
            except ValueError:
                continue

            if value.is_integer():
                value = int(value)
            sells_values.append(value)
    except:
        pass

    if len(sells_values) >= 2:
        print(f"✅ Продажи найдены через: {primary_selector}")
        print(f"📊 Продажи (второе значение): {sells_values[1]} шт.")
        return sells_values[1]

    if len(sells_values) == 1:
        print(f"✅ Продажи найдены через: {primary_selector}")
        print(f"📊 Продажи: {sells_values[0]} шт.")
        return sells_values[0]

    print("⚠️ Продажи не найдены")
    return 0


def find_stock(driver) -> int:
    """Поиск остатков товара"""
    print("\n📦 Шаг 6: Поиск остатков...")
    random_delay()

    stock_selectors = [
        "span.product-available__quantity",
        "span.quantity__value",
        "div.stock span",
        "[class*='stock'] span",
        "[class*='quantity']"
    ]

    for selector in stock_selectors:
        try:
            element = driver.find_element(By.CSS_SELECTOR, selector)
            text = element.text.strip()
            numbers = re.findall(r'\d+', text)
            if numbers:
                stock = int(numbers[0])
                print(f"✅ Остаток найден: {stock} шт.")
                return stock
        except:
            continue

    print("⚠️ Информация об остатках не найдена")
    return 0


# ============================================
# ПОДКЛЮЧЕНИЕ К БАЗЕ ДАННЫХ
# ============================================
def connect_db():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        logger.info("✅ Подключение к БД установлено")
        return conn
    except Exception as e:
        logger.error(f"❌ Ошибка подключения к БД: {e}")
        return None


def get_links_from_db() -> list[str]:
    """Получает список ссылок из public.reference_wildberries(link)."""
    conn = connect_db()
    if not conn:
        return []

    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT link
            FROM public.reference_wildberries
            WHERE link IS NOT NULL
              AND link <> ''
            ORDER BY link
            """
        )
        rows = cursor.fetchall()
        links = [row[0].strip() for row in rows if row[0] and row[0].strip()]
        logger.info(f"✅ Получено ссылок из БД: {len(links)}")

        # Выводим первые 5 ссылок для проверки
        print(f"\n🔗 ПЕРВЫЕ 5 ССЫЛОК ИЗ БД:")
        for i, link in enumerate(links[:5], 1):
            print(f"   {i}. {link}")

        return links
    except Exception as e:
        logger.error(f"❌ Ошибка чтения ссылок из public.reference_wildberries: {e}")
        return []
    finally:
        cursor.close()
        conn.close()


def get_previous_sales(conn, url: str) -> int:
    """Получает предыдущее значение sales_count для расчета daily_sales"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
                       SELECT sales_count
                       FROM public.wildberries
                       WHERE url = %s
                       ORDER BY parse_date DESC LIMIT 1
                       """, (url,))
        result = cursor.fetchone()
        cursor.close()
        return result[0] if result else 0
    except:
        return 0


# ============================================
# СОХРАНЕНИЕ ДАННЫХ В ТАБЛИЦУ (ДОЗАПИСЬ ВЕРСИЯ - ДИАГНОСТИКА)
# ============================================
def save_to_wildberries_table(conn, data: Dict) -> bool:
    """Сохранение данных в таблицу wildberries - КАЖДЫЙ РАЗ НОВАЯ СТРОКА"""
    try:
        cursor = conn.cursor()
        current_datetime = datetime.now()

        # Получаем данные из парсинга
        url = data.get('url')
        if not url:
            logger.error("❌ URL отсутствует в данных!")
            return False

        sales_count = data.get('sells', 0)
        regular_price = data.get('price', 0)
        old_price = data.get('old_price', 0)
        stock = data.get('stock', 0)
        artikul = data.get('vendor_code', 0)
        name = data.get('name', '')

        print(f"\n📝 ДАННЫЕ ДЛЯ СОХРАНЕНИЯ:")
        print(f"   Артикул: {artikul}")
        print(f"   Наименование: {name[:50]}")
        print(f"   Продажи: {sales_count}")
        print(f"   Цена: {regular_price}")

        # Получаем предыдущее значение sales_count для расчета daily_sales
        previous_sales = get_previous_sales(conn, url)
        daily_sales = max(0, sales_count - previous_sales)

        # Рассчитываем revenue (выручка)
        revenue = sales_count * regular_price

        print(f"\n📊 РАССЧИТАННЫЕ ПОКАЗАТЕЛИ:")
        print(f"   Предыдущие продажи: {previous_sales}")
        print(f"   Продажи за день: {daily_sales}")
        print(f"   Выручка: {revenue:.2f}")

        # ✅ ВСЕГДА ВСТАВЛЯЕМ НОВУЮ СТРОКУ, НЕ ОБНОВЛЯЯ СТАРЫЕ
        insert_query = """
                       INSERT INTO public.wildberries (url, \
                                                       artikul, \
                                                       name, \
                                                       parse_date, \
                                                       sales_count, \
                                                       daily_sales, \
                                                       revenue, \
                                                       stock, \
                                                       created_at, \
                                                       updated_at, \
                                                       regular_price, \
                                                       old_price) \
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) \
                       """

        values = (
            url,
            artikul,
            name,
            current_datetime,  # parse_date - дата этого парсинга
            sales_count,  # sales_count - общее количество продаж на текущий момент
            daily_sales,  # daily_sales - продажи за день (разница)
            revenue,  # revenue - выручка
            stock,  # stock - текущий остаток
            current_datetime,  # created_at
            current_datetime,  # updated_at
            regular_price,  # regular_price - текущая цена
            old_price  # old_price - старая цена (со скидкой)
        )

        print(f"\n💾 ВЫПОЛНЯЕМ INSERT...")
        cursor.execute(insert_query, values)
        conn.commit()

        # Проверяем, что запись действительно добавилась
        cursor.execute("SELECT COUNT(*) FROM public.wildberries WHERE artikul = %s AND parse_date = %s",
                       (artikul, current_datetime))
        count = cursor.fetchone()[0]

        if count > 0:
            logger.info(f"✅ Новая запись ДОБАВЛЕНА: {url} (артикул: {artikul})")
            print(f"✅ ПОДТВЕРЖДЕНО: Запись добавлена")
        else:
            logger.warning(f"⚠️ Запись не найдена после INSERT!")
            print(f"⚠️ ВНИМАНИЕ: Запись не найдена после вставки!")

        # Выводим информацию для проверки
        print(f"\n💾 СОХРАНЕНО В БД (НОВАЯ СТРОКА):")
        print(f"   Дата парсинга: {current_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"   URL: {url}")
        print(f"   Артикул: {artikul}")
        print(f"   Наименование: {name[:50] if name else 'Нет'}...")
        print(f"   Продажи всего (sales_count): {sales_count}")
        print(f"   Продажи за день (daily_sales): {daily_sales}")
        print(f"   Выручка (revenue): {revenue:.2f}")
        print(f"   Остаток (stock): {stock}")
        print(f"   Текущая цена (regular_price): {regular_price:.2f}")
        print(f"   Старая цена (old_price): {old_price:.2f}")

        cursor.close()
        return True

    except Exception as e:
        logger.error(f"❌ Ошибка сохранения в БД: {e}")
        print(f"❌ ДЕТАЛИ ОШИБКИ: {str(e)}")
        if conn:
            conn.rollback()
        return False


def save_to_excel(data: Dict) -> bool:
    """Сохраняет результат парсинга в Excel файл."""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "wildberries_data"
            ws.append(["artikul", "name", "url", "price", "old_price", "sells", "stock", "parse_date"])

        row_values = [
            data.get('vendor_code'),
            data.get('name', ''),
            data.get('url', ''),
            data.get('price', 0),
            data.get('old_price', 0),
            data.get('sells', 0),
            data.get('stock', 0),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]

        ws.append(row_values)
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"❌ Ошибка сохранения в Excel: {e}")
        return False


# ============================================
# ОСНОВНАЯ ФУНКЦИЯ ПАРСИНГА
# ============================================
def parse_wildberries(driver, url: str) -> Dict:
    """Основная функция парсинга товара с Wildberries"""
    print("\n" + "=" * 80)
    print(f"🔍 НАЧАЛО ПАРСИНГА: {url}")
    print("=" * 80)

    result = {
        'success': False,
        'url': url,
        'timestamp': datetime.now().isoformat()
    }

    try:
        logger.info(f"🌐 Открываем: {url}")
        driver.get(url)
        logger.info("⏳ Ожидаем загрузки страницы...")
        random_delay()

        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        logger.info("✅ Страница загружена")
        random_delay()

        print("\n" + "=" * 60)
        print("🔍 ПАРСИНГ ДАННЫХ ТОВАРА")
        print("=" * 60)

        # Парсинг всех данных
        result['price'] = find_current_price(driver)
        result['old_price'] = find_old_price(driver)
        result['sells'] = find_sells(driver)
        result['name'] = find_product_name(driver)
        result['vendor_code'], method = find_article_number(driver)
        result['stock'] = find_stock(driver)

        result['success'] = True

        # Вывод итоговой информации
        print("\n" + "=" * 60)
        print("📦 ИТОГОВЫЕ ДАННЫЕ")
        print("=" * 60)
        print(f"🔗 URL: {result['url']}")
        print(f"📝 Наименование: {result['name']}")
        print(f"🔢 Артикул: {result['vendor_code']} (найден через: {method})")
        print(f"💵 Обычная цена: {result['price']:.2f} ₽")
        print(f"💸 Зачеркнутая цена: {result['old_price']:.2f} ₽")
        print(f"📊 Продажи: {result['sells']} шт.")
        print(f"📦 Остаток: {result['stock']} шт.")

    except Exception as e:
        logger.error(f"❌ Критическая ошибка: {e}", exc_info=True)
        result['error'] = str(e)

    return result


# ============================================
# ФУНКЦИЯ ДЛЯ ПАРСИНГА НЕСКОЛЬКИХ ССЫЛОК
# ============================================
def parse_multiple_urls(driver, urls_list):
    """Парсинг нескольких ссылок (от 1 до 10)"""
    all_results = []

    for i, url in enumerate(urls_list, 1):
        print("\n" + "=" * 80)
        print(f"🔄 ПАРСИНГ ССЫЛКИ {i} ИЗ {len(urls_list)}")
        print("=" * 80)

        result = parse_wildberries(driver, url)
        all_results.append(result)

        # Пауза между ссылками
        if i < len(urls_list):
            print(f"\n⏸️ Пауза перед следующей ссылкой...")
            time.sleep(3)

    return all_results


# ============================================
# ЗАПУСК С СОХРАНЕНИЕМ В БД
# ============================================
def main():
    print("=" * 60)
    print("🚀 ЗАПУСК ПАРСЕРА WILDBERRIES")
    print("=" * 60)
    print(f"⚙️ Настройки задержек: от {MIN_DELAY} до {MAX_DELAY} секунд")
    print(f"📊 Таблица: {DB_SCHEMA}.{TABLE_NAME}")
    print("=" * 60)

    print(f"🔗 Источник ссылок: {DB_SCHEMA}.{REFERENCE_TABLE}(link)")
    print(f"📄 Excel файл: {EXCEL_FILE}")

    urls = get_links_from_db()
    if not urls:
        print(f"❌ В таблице {DB_SCHEMA}.{REFERENCE_TABLE} нет ссылок для парсинга.")
        return

    print(f"\n📋 Всего ссылок из БД для парсинга: {len(urls)}")

    print("\n🔗 Проверяем доступность Chrome debug endpoint...")
    if not is_debug_chrome_ready(CHROME_DEBUG_ADDRESS):
        print("❌ Chrome debug endpoint недоступен.")
        print("Запустите Chrome заранее по инструкции и повторите запуск парсера.")
        print("Ожидаемый endpoint: http://127.0.0.1:9222/json/version")
        return

    driver = setup_driver()
    try:
        # Парсинг
        start_time = time.time()
        all_results = parse_multiple_urls(driver, urls)
        duration = time.time() - start_time

        # Вывод статистики
        print("\n" + "=" * 60)
        print("📊 СТАТИСТИКА ПАРСИНГА")
        print("=" * 60)

        successful = sum(1 for r in all_results if r['success'])
        failed = len(all_results) - successful

        print(f"✅ Успешно: {successful} из {len(all_results)}")
        print(f"❌ Ошибок: {failed}")
        print(f"⏱️ Общее время: {duration:.2f} секунд")

        # Детальные результаты
        print("\n" + "=" * 60)
        print("📋 ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ")
        print("=" * 60)

        for i, result in enumerate(all_results, 1):
            print(f"\n📦 ТОВАР {i}:")
            print(f"   URL: {result['url']}")
            if result['success']:
                print(f"   ✅ Статус: Успешно")
                print(f"   📝 Наименование: {result['name']}")
                print(f"   🔢 Артикул: {result['vendor_code']}")
                print(f"   💵 Обычная цена: {result['price']:.2f} ₽")
                print(f"   💸 Зачеркнутая цена: {result['old_price']:.2f} ₽")
                print(f"   📊 Продажи: {result['sells']} шт.")
                print(f"   📦 Остаток: {result['stock']} шт.")
            else:
                print(f"   ❌ Статус: Ошибка")
                print(f"   Ошибка: {result.get('error', 'Неизвестная ошибка')}")

        # Сохранение в БД и Excel
        if successful > 0:
            conn = connect_db()
            saved_count = 0
            excel_saved_count = 0

            if conn:
                # Дополнительная проверка количества записей ДО сохранения
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM public.wildberries")
                before_count = cursor.fetchone()[0]
                print(f"\n📊 Записей в таблице ДО сохранения: {before_count}")
                cursor.close()

                for result in all_results:
                    if result['success']:
                        if save_to_wildberries_table(conn, result):
                            saved_count += 1
                        if save_to_excel(result):
                            excel_saved_count += 1

                # Проверка количества записей ПОСЛЕ сохранения
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM public.wildberries")
                after_count = cursor.fetchone()[0]
                print(f"\n📊 Записей в таблице ПОСЛЕ сохранения: {after_count}")
                print(f"📊 Добавлено новых записей: {after_count - before_count}")
                cursor.close()

                conn.close()
                print(f"\n✅ Сохранено в БД {DB_SCHEMA}.{TABLE_NAME}: {saved_count} из {successful} товаров")
                print(f"✅ Сохранено в Excel {EXCEL_FILE}: {excel_saved_count} из {successful} товаров")
            else:
                print("⚠️ Нет подключения к БД. Сохранение в таблицу пропущено.")
                for result in all_results:
                    if result['success'] and save_to_excel(result):
                        excel_saved_count += 1
                print(f"✅ Сохранено в Excel {EXCEL_FILE}: {excel_saved_count} из {successful} товаров")

        print("\n" + "=" * 60)
        print("🏁 РАБОТА ПАРСЕРА ЗАВЕРШЕНА")
        print("=" * 60)
    finally:
        print("\n🔄 Завершаем сессию Selenium...")
        short_delay()
        driver.quit()
        logger.info("✅ Selenium-сессия закрыта")


# ============================================
# ЗАПУСК
# ============================================
if __name__ == "__main__":
    main()
