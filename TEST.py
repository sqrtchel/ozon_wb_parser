import time
import re
import sys
import os
import shutil
import subprocess
import urllib.request
import psycopg2
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from datetime import datetime
from openpyxl import Workbook, load_workbook

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# ========== НАСТРОЙКИ POSTGRESQL ==========
DB_CONFIG = {
    'host': 'localhost',  # твой хост
    'port': 5432,  # порт PostgreSQL
    'database': 'postgres',  # название твоей БД
    'user': 'postgres',  # твой пользователь
    'password': 's56a12y89'  # твой пароль
}
EXCEL_FILE = r"C:\Users\user\OneDrive\Рабочий стол\парсер\results.xlsx"
CHROME_USER_DATA_DIR = os.environ.get(
    "CHROME_USER_DATA_DIR",
    r"C:\Users\user\AppData\Local\Google\Chrome\User Data"
)
CHROME_PROFILE_DIRECTORY = os.environ.get("CHROME_PROFILE_DIRECTORY", "Default")
AUTOMATION_USER_DATA_DIR = os.environ.get(
    "AUTOMATION_USER_DATA_DIR",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "chrome_automation_user_data")
)
MPSTATS_EXTENSION_ID = os.environ.get("MPSTATS_EXTENSION_ID", "pjbepnginjokklnhdgladnmlghcchbeb")
FORCE_MPSTATS_SYNC = os.environ.get("FORCE_MPSTATS_SYNC", "1") == "1"
CHROME_DEBUG_ADDRESS = os.environ.get("CHROME_DEBUG_ADDRESS", "127.0.0.1:9222")
CHROME_BINARY = os.environ.get("CHROME_BINARY", r"C:\Program Files\Google\Chrome\Application\chrome.exe")
SELENIUM_PROFILE_DIR = os.environ.get(
    "SELENIUM_PROFILE_DIR",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "selenium_profile")
)
SELENIUM_PROFILE_NAME = os.environ.get("SELENIUM_PROFILE_NAME", "Default")
FIRST_RUN_AUTH_WAIT_SEC = int(os.environ.get("FIRST_RUN_AUTH_WAIT_SEC", "90"))


def _kill_chrome_processes():
    """Закрывает фоновые процессы Chrome перед стартом WebDriver."""
    try:
        subprocess.run(
            ["taskkill", "/F", "/IM", "chrome.exe", "/T"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False
        )
        time.sleep(1.2)
    except Exception:
        pass


def _get_chrome_binary():
    if os.path.exists(CHROME_BINARY):
        return CHROME_BINARY
    fallback = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    if os.path.exists(fallback):
        return fallback
    return None


def _is_debug_chrome_ready(address=None):
    target_address = address or CHROME_DEBUG_ADDRESS
    try:
        urllib.request.urlopen(f"http://{target_address}/json/version", timeout=1.0)
        return True
    except Exception:
        return False


def _get_available_debug_address():
    """
    Возвращает первый доступный debug endpoint Chrome.
    Проверяет заданный CHROME_DEBUG_ADDRESS и стандартные fallback-варианты.
    """
    candidates = [CHROME_DEBUG_ADDRESS, "127.0.0.1:9222", "localhost:9222", "127.0.0.1:9223", "localhost:9223", "127.0.0.1:9333", "localhost:9333"]
    seen = set()
    for address in candidates:
        if address in seen:
            continue
        seen.add(address)
        if _is_debug_chrome_ready(address):
            return address
    return None


def _start_debug_chrome():
    global CHROME_DEBUG_ADDRESS

    chrome_binary = _get_chrome_binary()
    if not chrome_binary:
        raise RuntimeError("Не найден chrome.exe. Укажите путь через переменную CHROME_BINARY.")

    host = CHROME_DEBUG_ADDRESS.split(":")[0]
    preferred_port = CHROME_DEBUG_ADDRESS.split(":")[-1]
    candidate_ports = [preferred_port, "9223", "9333"]

    for port in candidate_ports:
        target_address = f"{host}:{port}"
        cmd = [
            chrome_binary,
            f"--remote-debugging-port={port}",
            f"--remote-debugging-address={host}",
            "--remote-allow-origins=*",
            f"--user-data-dir={CHROME_USER_DATA_DIR}",
            f"--profile-directory={CHROME_PROFILE_DIRECTORY}",
            "--no-first-run",
            "--no-default-browser-check",
            "about:blank"
        ]

        proc = subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        for _ in range(120):  # до ~60 сек на тяжёлый профиль
            if _is_debug_chrome_ready(target_address):
                CHROME_DEBUG_ADDRESS = target_address
                return
            if proc.poll() is not None:
                break
            time.sleep(0.5)

    raise RuntimeError("Chrome запущен, но debug endpoint не поднялся ни на одном порту.")


def _cleanup_chrome_profile_locks(user_data_dir):
    """Удаляет lock-файлы Chrome, мешающие запуску профиля."""
    lock_files = ["SingletonLock", "SingletonCookie", "SingletonSocket"]
    for lock_name in lock_files:
        lock_path = os.path.join(user_data_dir, lock_name)
        try:
            if os.path.exists(lock_path):
                os.remove(lock_path)
        except Exception:
            pass


def _replace_dir(src_path, dst_path):
    """Перекопировать папку целиком (с удалением старой версии)."""
    if not os.path.exists(src_path):
        return
    if os.path.exists(dst_path):
        shutil.rmtree(dst_path, ignore_errors=True)
    os.makedirs(os.path.dirname(dst_path), exist_ok=True)
    shutil.copytree(src_path, dst_path)


def _copy_file_if_exists(src_path, dst_path):
    if not os.path.exists(src_path):
        return
    os.makedirs(os.path.dirname(dst_path), exist_ok=True)
    shutil.copy2(src_path, dst_path)


def _get_latest_extension_dir(user_data_dir, profile_directory, extension_id):
    ext_root = os.path.join(user_data_dir, profile_directory, "Extensions", extension_id)
    if not os.path.isdir(ext_root):
        return None
    versions = [
        name for name in os.listdir(ext_root)
        if os.path.isdir(os.path.join(ext_root, name))
    ]
    if not versions:
        return None
    versions.sort(reverse=True)
    return os.path.join(ext_root, versions[0])


def _sync_mpstats_profile(force=False):
    """
    Создаёт лёгкий automation-профиль и переносит в него только данные mpstats.
    Так Chrome стартует стабильнее, но расширение остаётся рабочим.
    """
    src_profile = os.path.join(CHROME_USER_DATA_DIR, CHROME_PROFILE_DIRECTORY)
    dst_profile = os.path.join(AUTOMATION_USER_DATA_DIR, "Default")
    os.makedirs(dst_profile, exist_ok=True)

    marker_path = os.path.join(dst_profile, ".mpstats_profile_ready")
    if os.path.exists(marker_path) and not force:
        print("  ♻️ Используем существующий automation-профиль (с сохраненной авторизацией).")
        return

    # Базовые файлы профиля
    _copy_file_if_exists(
        os.path.join(CHROME_USER_DATA_DIR, "Local State"),
        os.path.join(AUTOMATION_USER_DATA_DIR, "Local State")
    )
    _copy_file_if_exists(
        os.path.join(src_profile, "Preferences"),
        os.path.join(dst_profile, "Preferences")
    )
    _copy_file_if_exists(
        os.path.join(src_profile, "Secure Preferences"),
        os.path.join(dst_profile, "Secure Preferences")
    )

    # Ключевые каталоги расширения mpstats
    rel_dirs = [
        os.path.join("Extensions", MPSTATS_EXTENSION_ID),
        os.path.join("Local Extension Settings", MPSTATS_EXTENSION_ID),
        os.path.join("Sync Extension Settings", MPSTATS_EXTENSION_ID),
        os.path.join("Extension State"),
        os.path.join("Extension Rules"),
        os.path.join("Extension Scripts"),
        os.path.join("IndexedDB", "https_mpstats.io_0.indexeddb.leveldb"),
        os.path.join("Local Storage", "leveldb"),
        os.path.join("Session Storage"),
        os.path.join("WebStorage")
    ]

    for rel_dir in rel_dirs:
        _replace_dir(
            os.path.join(src_profile, rel_dir),
            os.path.join(dst_profile, rel_dir)
        )

    # Cookies часто нужны расширению для авторизации/запросов.
    _copy_file_if_exists(
        os.path.join(src_profile, "Network", "Cookies"),
        os.path.join(dst_profile, "Network", "Cookies")
    )
    _copy_file_if_exists(
        os.path.join(src_profile, "Network", "Cookies-journal"),
        os.path.join(dst_profile, "Network", "Cookies-journal")
    )

    _copy_file_if_exists(
        os.path.join(src_profile, "Login Data"),
        os.path.join(dst_profile, "Login Data")
    )

    # Маркер, что первичная синхронизация завершена.
    with open(marker_path, "w", encoding="utf-8") as marker_file:
        marker_file.write("ready")


def _sync_mpstats_to_selenium_profile():
    """Переносит данные расширения mpstats в постоянный selenium-профиль."""
    src_profile = os.path.join(CHROME_USER_DATA_DIR, CHROME_PROFILE_DIRECTORY)
    dst_profile = os.path.join(SELENIUM_PROFILE_DIR, SELENIUM_PROFILE_NAME)
    os.makedirs(dst_profile, exist_ok=True)

    rel_dirs = [
        os.path.join("Extensions", MPSTATS_EXTENSION_ID),
        os.path.join("Local Extension Settings", MPSTATS_EXTENSION_ID),
        os.path.join("Sync Extension Settings", MPSTATS_EXTENSION_ID),
        os.path.join("Extension State"),
        os.path.join("Extension Rules"),
        os.path.join("Extension Scripts"),
        os.path.join("IndexedDB", "https_mpstats.io_0.indexeddb.leveldb"),
        os.path.join("Local Storage", "leveldb"),
        os.path.join("Session Storage"),
        os.path.join("WebStorage"),
    ]

    for rel_dir in rel_dirs:
        _replace_dir(
            os.path.join(src_profile, rel_dir),
            os.path.join(dst_profile, rel_dir)
        )

    _copy_file_if_exists(
        os.path.join(src_profile, "Preferences"),
        os.path.join(dst_profile, "Preferences")
    )
    _copy_file_if_exists(
        os.path.join(src_profile, "Secure Preferences"),
        os.path.join(dst_profile, "Secure Preferences")
    )
    _copy_file_if_exists(
        os.path.join(src_profile, "Network", "Cookies"),
        os.path.join(dst_profile, "Network", "Cookies")
    )
    _copy_file_if_exists(
        os.path.join(src_profile, "Network", "Cookies-journal"),
        os.path.join(dst_profile, "Network", "Cookies-journal")
    )


def get_db_connection():
    """Создаёт соединение с PostgreSQL"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"❌ Ошибка подключения к БД: {e}")
        return None


def get_links_from_db():
    """Получает список ссылок из public.reference_ozon(link)"""
    conn = get_db_connection()
    if not conn:
        return []

    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT link
            FROM public.reference_ozon
            WHERE link IS NOT NULL
              AND link <> ''
            ORDER BY link
        """)
        rows = cursor.fetchall()
        return [row[0].strip() for row in rows if row[0] and row[0].strip()]
    except Exception as e:
        print(f"❌ Ошибка чтения ссылок из public.reference_ozon: {e}")
        return []
    finally:
        cursor.close()
        conn.close()


def save_to_db(artikul, name, le, price, old_price, sells, remaining_product, data_text):
    """Сохраняет данные в таблицу ozon"""
    conn = get_db_connection()
    if not conn:
        return False

    cursor = conn.cursor()

    try:
        if artikul:
            # Всегда вставляем новую запись (без обновления старых).
            cursor.execute('''
                INSERT INTO ozon 
                (artikul, name, le, price, old_price, sells, remaining_product, data)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ''', (artikul, name, le, price, old_price, sells, remaining_product, data_text))
            print(f"    ✅ Добавлена новая запись для артикула {artikul}")
        else:
            # Вставляем без артикула
            cursor.execute('''
                INSERT INTO ozon 
                (name, le, price, old_price, sells, remaining_product, data)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (name, le, price, old_price, sells, remaining_product, data_text))
            print(f"    ✅ Добавлена новая запись для товара: {name[:50]}")

        conn.commit()
        return True

    except Exception as e:
        print(f"    ❌ Ошибка сохранения в БД: {e}")
        conn.rollback()
        return False
    finally:
        cursor.close()
        conn.close()


def save_to_excel(artikul, name, le, price, old_price, sells, remaining_product, data_text):
    """Сохраняет данные в Excel файл (.xlsx)"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "ozon_data"
            ws.append([
                "artikul",
                "name",
                "le",
                "price",
                "old_price",
                "sells",
                "remaining_product",
                "data"
            ])

        row_values = [artikul, name, le, price, old_price, sells, remaining_product, data_text]

        # Ищем последнюю реально заполненную строку, чтобы игнорировать пустые строки внизу листа.
        last_filled_row = 0
        for row_idx in range(ws.max_row, 0, -1):
            row_has_data = any(
                ws.cell(row=row_idx, column=col_idx).value not in (None, "")
                for col_idx in range(1, len(row_values) + 1)
            )
            if row_has_data:
                last_filled_row = row_idx
                break

        target_row = last_filled_row + 1
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col_idx, value=value)

        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"    ❌ Ошибка сохранения в Excel: {e}")
        return False


def setup_driver():
    """
    Подключение Selenium к уже открытому Chrome через debug-порт.
    Chrome должен быть запущен пользователем заранее.
    """
    try:
        available_address = _get_available_debug_address()
        if not available_address:
            raise RuntimeError(
                f"Chrome debug endpoint недоступен: {CHROME_DEBUG_ADDRESS}.\n"
                "Запустите Chrome заранее командой:\n"
                "\"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe\" "
                "--remote-debugging-port=9222 "
                "--remote-debugging-address=127.0.0.1 "
                "--user-data-dir=\"C:\\Users\\user\\AppData\\Local\\Google\\Chrome\\User Data\" "
                "--profile-directory=Default"
            )

        print(f"  🔗 Подключаемся к открытому Chrome: {available_address}")
        options = Options()
        options.debugger_address = available_address

        driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(40)
        print("  ✅ Подключение к Chrome успешно.")
        return driver
    except Exception as e:
        raise RuntimeError(
            "Не удалось подключиться к Chrome через debug-порт. "
            "Проверьте, что профиль Default не заблокирован другим процессом."
        ) from e


def parse_ozon_product(driver, url):
    """Парсит информацию о товаре с Ozon"""
    try:
        print(f"\n🌐 Загружаем страницу: {url}")
        try:
            driver.get(url)
        except TimeoutException:
            print("  ⚠️ Долгая загрузка страницы, продолжаем парсинг по уже загруженному контенту...")
            try:
                driver.execute_script("window.stop();")
            except Exception:
                pass
        time.sleep(5)  # Ждём загрузки основной страницы
        time.sleep(3)  # Дополнительная задержка

        product_data = {
            'artikul': None,
            'name': None,
            'price': None,
            'old_price': None,
            'sells': None,
            'remaining': None,
            'le': url
        }

        # 1. Парсим название товара
        print("  🔍 Ищем название...")
        name_selectors = [
            "div[data-widget='webProductHeading'] h1",
            "h1[itemprop='name']",
            "div[data-widget='webProductHeading'] span",
            "h1.tsHeadline550Medium"
        ]

        for selector in name_selectors:
            try:
                elem = driver.find_element(By.CSS_SELECTOR, selector)
                product_data['name'] = elem.text.strip()
                if product_data['name']:
                    print(f"    📝 Название: {product_data['name'][:60]}...")
                    break
            except:
                continue

        # 2. Парсим артикул (обновлённый селектор)
        print("  🔍 Ищем артикул...")
        artikul_selectors = [
            "div.ga5_3_15-a3.tsBodyControl400Small",  # ваш точный селектор
            "div[class*='ga5']",  # более широкий вариант
            "div.tsBodyControl400Small"  # по классу
        ]

        for selector in artikul_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    text = elem.text
                    if 'Артикул:' in text or 'Артикул' in text:
                        # Извлекаем число
                        artikul_match = re.search(r'Артикул:\s*(\d+)', text)
                        if artikul_match:
                            product_data['artikul'] = int(artikul_match.group(1))
                            print(f"    🔢 Артикул: {product_data['artikul']}")
                            break
                        else:
                            # Пробуем просто взять все цифры
                            digits = re.sub(r'[^\d]', '', text)
                            if digits:
                                product_data['artikul'] = int(digits)
                                print(f"    🔢 Артикул: {product_data['artikul']}")
                                break
                if product_data['artikul']:
                    break
            except:
                continue

        # Если не нашли, пробуем через XPath
        if not product_data['artikul']:
            try:
                artikul_elem = driver.find_element(By.XPATH, "//div[contains(text(), 'Артикул:')]")
                text = artikul_elem.text
                artikul_match = re.search(r'(\d+)', text)
                if artikul_match:
                    product_data['artikul'] = int(artikul_match.group(1))
                    print(f"    🔢 Артикул: {product_data['artikul']}")
            except:
                pass

        # 3. Парсим цену (обновлённый селектор)
        print("  🔍 Ищем цену...")
        price_selectors = [
            "span.pdp_bj.tsHeadline500Medium",  # новый приоритетный селектор
            "span.pdp_bj",
            "span[class*='pdp_i9b']",  # более широкий вариант
            "span.tsBody400Small",  # по классу
            "div[data-widget='webPrice'] span[class*='final']",
            "div[data-widget='webPrice'] span"
        ]

        for selector in price_selectors:
            try:
                elem = driver.find_element(By.CSS_SELECTOR, selector)
                text = elem.text
                # Удаляем пробелы и символы, оставляем только цифры
                clean = re.sub(r'[^\d]', '', text)
                if clean and clean.isdigit():
                    price = int(clean)
                    if 100 < price < 500000:
                        product_data['price'] = price
                        print(f"    💰 Цена: {product_data['price']} ₽")
                        break
            except:
                continue

        # 3.1 Парсим зачеркнутую (старую) цену
        print("  🔍 Ищем старую цену...")
        old_price_selectors = [
            "span.pdp_i9b.pdp_bj0.pdp_bi9.tsBody400Small",
            "span.pdp_i9b.pdp_bj0",
            "span.pdp_i9b"
        ]

        for selector in old_price_selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    text = elem.text
                    clean = re.sub(r'[^\d]', '', text)
                    if not clean or not clean.isdigit():
                        continue

                    value = int(clean)
                    if 100 < value < 1000000:
                        product_data['old_price'] = value
                        print(f"    🏷️ Старая цена: {product_data['old_price']} ₽")
                        break
                if product_data['old_price'] is not None:
                    break
            except:
                continue

        # 4. Парсим количество продаж и остаток
        print("  🔍 Ищем продажи и остаток...")
        try:
            # Приоритетный поиск продаж по новому span:
            # <span class="_text_j8xb1_1 _body-md-medium_j8xb1_75 _default_j8xb1_109">61,2 шт. </span>
            if product_data['sells'] is None:
                new_sells_selectors = [
                    "span._text_j8xb1_1._body-md-medium_j8xb1_75._default_j8xb1_109",
                    "span[class*='_body-md-medium_j8xb1_75'][class*='_default_j8xb1_109']",
                    "span[class*='_text_j8xb1_1'][class*='_body-md-medium_j8xb1_75']"
                ]
                for selector in new_sells_selectors:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for elem in elements:
                        text = elem.text.strip().lower()
                        if 'шт' not in text:
                            continue

                        sells_match = re.search(r'(\d+(?:[ \u00a0]\d{3})*(?:[.,]\d+)?)\s*шт', text)
                        if not sells_match:
                            continue

                        normalized = sells_match.group(1).replace(' ', '').replace('\u00a0', '').replace(',', '.')
                        try:
                            sells_value = float(normalized)
                            if sells_value.is_integer():
                                sells_value = int(sells_value)
                            product_data['sells'] = sells_value
                            print(f"    📊 Продажи: {product_data['sells']} шт.")
                            break
                        except ValueError:
                            continue
                    if product_data['sells'] is not None:
                        break

            # Ищем блоки с информацией о продажах
            sell_selectors = [
                "div[data-widget='webProductHeading'] div span",
                "div.tsBodyControl400Small"
            ]

            for selector in sell_selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    text = elem.text.lower()
                    if 'продано' in text:
                        sells_text = re.sub(r'[^\d]', '', text)
                        if sells_text:
                            product_data['sells'] = int(sells_text)
                            print(f"    📊 Продано: {product_data['sells']} шт.")
                    elif 'осталось' in text or 'остаток' in text:
                        remaining_text = re.sub(r'[^\d]', '', text)
                        if remaining_text:
                            product_data['remaining'] = int(remaining_text)
                            print(f"    📦 Остаток: {product_data['remaining']} шт.")

            # Отдельный точный поиск остатка:
            # <span class="tsCompactControl300XSmall" style="color:var(--textAccent);">3</span>
            if product_data['remaining'] is None:
                remaining_selectors = [
                    "span.tsCompactControl300XSmall[style*='--textAccent']",
                    "span[class*='tsCompactControl300XSmall'][style*='--textAccent']",
                    "span.tsCompactControl300XSmall"
                ]
                fallback_remaining = None
                for selector in remaining_selectors:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for elem in elements:
                        text = elem.text.strip()
                        remaining_match = re.search(r'^\d+$', text)
                        if not remaining_match:
                            continue

                        value = int(remaining_match.group(0))
                        parent_text = ""
                        try:
                            parent_text = elem.find_element(By.XPATH, "./..").text.lower()
                        except:
                            pass

                        if 'остал' in parent_text or 'единиц' in parent_text:
                            product_data['remaining'] = value
                            print(f"    📦 Остаток: {product_data['remaining']} шт.")
                            break

                        # fallback: берем первое найденное число из нужного span
                        if fallback_remaining is None:
                            fallback_remaining = value
                    if product_data['remaining'] is not None:
                        break

                if product_data['remaining'] is None and fallback_remaining is not None:
                    product_data['remaining'] = fallback_remaining
                    print(f"    📦 Остаток: {product_data['remaining']} шт.")
        except:
            pass

        return product_data

    except Exception as e:
        print(f"  ❌ Ошибка при парсинге: {e}")
        return None


def main():
    """Основная функция: парсит ссылки из PostgreSQL"""
    print("=" * 60)
    print("🕷️  ПАРСЕР OZON (Режим: ссылки из PostgreSQL)")
    print("=" * 60)

    urls = get_links_from_db()
    if not urls:
        print("❌ В таблице public.reference_ozon нет ссылок для парсинга.")
        return

    print(f"\n📌 Найдено ссылок в БД: {len(urls)}")

    # Запускаем браузер
    print("\n🚀 Запускаем Chrome...")
    driver = setup_driver()

    try:
        for i, url in enumerate(urls, start=1):
            print("\n" + "=" * 60)
            print(f"🔗 [{i}/{len(urls)}] Ссылка: {url}")

            if not url.startswith('http'):
                print("❌ Неверный формат ссылки. Ссылка должна начинаться с http:// или https://")
                continue

            # Парсим товар
            product_data = parse_ozon_product(driver, url)

            if product_data and product_data.get('price'):
                print("\n" + "=" * 60)
                print("📦 РЕЗУЛЬТАТ ПАРСИНГА:")
                print(f"  📝 Название: {product_data.get('name', 'Не найдено')}")
                print(f"  🔢 Артикул: {product_data.get('artikul', 'Не найден')}")
                print(f"  💰 Цена: {product_data.get('price', 'Не найдена')} ₽")
                print(f"  🏷️ Старая цена: {product_data.get('old_price', 'Нет данных')} ₽")
                print(f"  📊 Продажи: {product_data.get('sells', 'Нет данных')} шт.")
                print(f"  📦 Остаток: {product_data.get('remaining', 'Нет данных')} шт.")
                print("=" * 60)

                # Сохраняем в БД
                print("\n💾 Сохраняем в PostgreSQL...")
                saved = save_to_db(
                    artikul=product_data.get('artikul'),
                    name=product_data.get('name', 'Неизвестный товар'),
                    le=url,
                    price=product_data.get('price'),
                    old_price=product_data.get('old_price'),
                    sells=product_data.get('sells', 0),
                    remaining_product=product_data.get('remaining', 0),
                    data_text=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                if saved:
                    print("✅ Данные сохранены!")
                else:
                    print("❌ Не удалось сохранить данные в таблицу ozon")

                print("📄 Сохраняем в Excel...")
                excel_saved = save_to_excel(
                    artikul=product_data.get('artikul'),
                    name=product_data.get('name', 'Неизвестный товар'),
                    le=url,
                    price=product_data.get('price'),
                    old_price=product_data.get('old_price'),
                    sells=product_data.get('sells', 0),
                    remaining_product=product_data.get('remaining', 0),
                    data_text=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                if excel_saved:
                    print(f"✅ Данные сохранены в {EXCEL_FILE}")
                else:
                    print("❌ Не удалось сохранить данные в Excel файл")
            else:
                print("\n❌ Не удалось получить цену товара. Возможно, страница заблокирована или изменилась структура.")

            print("\n⏳ Ждём 3 секунды перед следующей ссылкой...")
            time.sleep(3)

    except KeyboardInterrupt:
        print("\n⚠️ Программа прервана пользователем")
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
    finally:
        driver.quit()
        print("\n🚪 Браузер закрыт")
        print("👋 До свидания!")


if __name__ == "__main__":
    main()